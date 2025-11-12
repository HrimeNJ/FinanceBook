"""
ExportModule for exporting data to JSON/Excel formats
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExportModule:
    """数据导出模块"""

    def __init__(self, state):
        """初始化导出模块

        Args:
            state: AppState对象,包含用户数据
        """
        self.state = state
        self.export_dir = Path.cwd() / "FinanceBookExports"
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export_to_json(self, filename: Optional[str] = None) -> tuple[bool, str]:
        """导出数据为JSON格式

        Args:
            filename: 文件名(可选),默认使用时间戳

        Returns:
            (成功标志, 文件路径或错误信息)
        """
        try:
            if not self.state.current_user:
                return False, "未登录用户"

            # 生成文件名
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"finance_data_{timestamp}.json"

            if not filename.endswith(".json"):
                filename += ".json"

            filepath = self.export_dir / filename

            # 准备导出数据
            export_data = {
                "export_time": datetime.now().isoformat(),
                "user": {
                    "user_id": self.state.current_user.user_id,
                    "username": self.state.current_user.username,
                    "email": self.state.current_user.email,
                },
                "categories": [
                    {
                        "category_id": cat.category_id,
                        "name": cat.name,
                    }
                    for cat in self.state.categories
                ],
                "records": [
                    {
                        "record_id": record.record_id,
                        "amount": record.amount,
                        "date": record.date.isoformat(),
                        "record_type": record.record_type,
                        "category_id": record.category_id,
                        "category_name": (
                            self.state.get_category_by_id(record.category_id).name
                            if self.state.get_category_by_id(record.category_id)
                            else "未知"
                        ),
                        "note": record.note,
                        "created_at": (
                            record.created_at.isoformat() if record.created_at else None
                        ),
                        "updated_at": (
                            record.updated_at.isoformat() if record.updated_at else None
                        ),
                    }
                    for record in self.state.records
                ],
                "summary": {
                    "total_records": len(self.state.records),
                    "total_income": sum(
                        r.amount
                        for r in self.state.records
                        if r.record_type == "income"
                    ),
                    "total_expense": sum(
                        r.amount
                        for r in self.state.records
                        if r.record_type == "expense"
                    ),
                },
            }

            # 写入文件
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)

            return True, str(filepath)

        except Exception as e:
            return False, f"导出失败: {str(e)}"

    def export_to_excel(self, filename: Optional[str] = None) -> tuple[bool, str]:
        """导出数据为Excel格式

        Args:
            filename: 文件名(可选),默认使用时间戳

        Returns:
            (成功标志, 文件路径或错误信息)
        """
        if not EXCEL_AVAILABLE:
            return False, "需要安装 openpyxl 库: pip install openpyxl"

        try:
            if not self.state.current_user:
                return False, "未登录用户"

            # 生成文件名
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"finance_data_{timestamp}.xlsx"

            if not filename.endswith(".xlsx"):
                filename += ".xlsx"

            filepath = self.export_dir / filename

            # 创建工作簿
            wb = openpyxl.Workbook()

            # 删除默认工作表
            wb.remove(wb.active)

            # 创建记录表
            self._create_records_sheet(wb)

            # 创建分类表
            self._create_categories_sheet(wb)

            # 创建汇总表
            self._create_summary_sheet(wb)

            # 保存文件
            wb.save(filepath)

            return True, str(filepath)

        except Exception as e:
            return False, f"导出失败: {str(e)}"

    def _create_records_sheet(self, wb):
        """创建记录工作表"""
        ws = wb.create_sheet("记账记录")

        # 设置表头
        headers = [
            "记录ID",
            "日期",
            "类型",
            "分类",
            "金额",
            "备注",
            "创建时间",
            "更新时间",
        ]
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填充数据
        for record in self.state.records:
            category = self.state.get_category_by_id(record.category_id)
            category_name = category.name if category else "未知"

            ws.append(
                [
                    record.record_id,
                    record.date.strftime("%Y-%m-%d"),
                    "收入" if record.record_type == "income" else "支出",
                    category_name,
                    record.amount,
                    record.note or "",
                    (
                        record.created_at.strftime("%Y-%m-%d %H:%M:%S")
                        if record.created_at
                        else ""
                    ),
                    (
                        record.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                        if record.updated_at
                        else ""
                    ),
                ]
            )

        # 调整列宽
        column_widths = [10, 12, 10, 15, 12, 30, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    def _create_categories_sheet(self, wb):
        """创建分类工作表"""
        ws = wb.create_sheet("分类列表")

        # 设置表头
        headers = ["分类ID", "分类名称", "类型", "图标", "颜色"]
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填充数据
        for category in self.state.categories:
            ws.append(
                [
                    category.category_id,
                    category.name,
                ]
            )

        # 调整列宽
        for i, width in enumerate([10, 15, 10, 15, 15], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    def _create_summary_sheet(self, wb):
        """创建汇总工作表"""
        ws = wb.create_sheet("数据汇总", 0)  # 设为第一个工作表

        # 用户信息
        ws.append(["用户信息"])
        ws.append(["用户名", self.state.current_user.username])
        ws.append(["邮箱", self.state.current_user.email])
        ws.append([])

        # 统计信息
        total_income = sum(
            r.amount for r in self.state.records if r.record_type == "income"
        )
        total_expense = sum(
            r.amount for r in self.state.records if r.record_type == "expense"
        )
        net_savings = total_income - total_expense

        ws.append(["财务统计"])
        ws.append(["总记录数", len(self.state.records)])
        ws.append(["总收入", f"¥{total_income:.2f}"])
        ws.append(["总支出", f"¥{total_expense:.2f}"])
        ws.append(["净储蓄", f"¥{net_savings:.2f}"])
        ws.append([])

        # 导出信息
        ws.append(["导出信息"])
        ws.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["数据版本", "1.0.0"])

        # 设置样式
        title_font = Font(bold=True, size=12)
        for row in [1, 5, 11]:
            ws[f"A{row}"].font = title_font

        # 调整列宽
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30

    def get_export_directory(self) -> str:
        """获取导出目录路径"""
        return str(self.export_dir)
